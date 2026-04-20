from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .utils import coerce_float


DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    )
}

ISHARES_DOWNLOAD_URL_TEMPLATE = (
    "https://www.ishares.com/us/products/{product_id}/"
    "ishares-{fund_slug}/1467271812596.ajax?fileType=csv&fileName={symbol}_holdings&dataType=fund"
)

SPDR_SECTOR_URLS = {
    "XLB": "https://www.ssga.com/us/en/intermediary/etfs/the-materials-select-sector-spdr-fund-xlb",
    "XLC": "https://www.ssga.com/us/en/intermediary/etfs/the-communication-services-select-sector-spdr-fund-xlc",
    "XLE": "https://www.ssga.com/us/en/intermediary/etfs/the-energy-select-sector-spdr-fund-xle",
    "XLF": "https://www.ssga.com/us/en/intermediary/etfs/the-financial-select-sector-spdr-fund-xlf",
    "XLI": "https://www.ssga.com/us/en/intermediary/etfs/the-industrial-select-sector-spdr-fund-xli",
    "XLK": "https://www.ssga.com/us/en/intermediary/etfs/the-technology-select-sector-spdr-fund-xlk",
    "XLP": "https://www.ssga.com/us/en/intermediary/etfs/the-consumer-staples-select-sector-spdr-fund-xlp",
    "XLRE": "https://www.ssga.com/us/en/intermediary/etfs/the-real-estate-select-sector-spdr-fund-xlre",
    "XLU": "https://www.ssga.com/us/en/intermediary/etfs/the-utilities-select-sector-spdr-fund-xlu",
    "XLV": "https://www.ssga.com/us/en/intermediary/etfs/the-health-care-select-sector-spdr-fund-xlv",
    "XLY": "https://www.ssga.com/us/en/intermediary/etfs/the-consumer-discretionary-select-sector-spdr-fund-xly",
}

SPDR_HOLDINGS_URL_TEMPLATES = [
    "https://www.ssga.com/library-content/products/fund-data/etfs/us/holdings-daily-us-en-{symbol}.xlsx",
    "https://www.ssga.com/us/en/intermediary/etfs/library-content/products/fund-data/etfs/us/holdings-daily-us-en-{symbol}.xlsx",
]


class HoldingsExportError(RuntimeError):
    """Raised when a holdings source cannot be exported."""


@dataclass(slots=True)
class NormalizedHolding:
    ticker: str
    issuer_name: str
    sector: Optional[str]
    weight: Optional[float]
    cik: Optional[str] = None
    raw_identifier: Optional[str] = None
    source_url: Optional[str] = None
    etf_symbol: Optional[str] = None

    def as_record(self) -> dict:
        payload = {
            "ticker": self.ticker,
            "issuer_name": self.issuer_name,
            "sector": self.sector,
            "weight": self.weight,
            "cik": self.cik,
        }
        if self.raw_identifier:
            payload["raw_identifier"] = self.raw_identifier
        if self.source_url:
            payload["source_url"] = self.source_url
        if self.etf_symbol:
            payload["etf_symbol"] = self.etf_symbol
        return payload


class HoldingsHtmlParser:
    COLUMN_ALIASES = {
        "ticker": {"ticker", "symbol", "holding ticker", "fund ticker"},
        "issuer_name": {
            "issuer name",
            "company name",
            "name",
            "security",
            "holding name",
            "description",
        },
        "sector": {"sector", "gics sector", "market sector"},
        "weight": {"weight", "index weight", "portfolio weight", "% of net assets", "fund weight"},
        "cik": {"cik"},
    }

    def parse(self, html: str, source_url: str, default_sector: Optional[str] = None, etf_symbol: Optional[str] = None) -> list[NormalizedHolding]:
        soup = BeautifulSoup(html, "html.parser")
        candidates: list[list[NormalizedHolding]] = []
        for table in soup.find_all("table"):
            parsed = self._parse_table(table, source_url=source_url, default_sector=default_sector, etf_symbol=etf_symbol)
            if parsed:
                candidates.append(parsed)
        if not candidates:
            raise HoldingsExportError(f"no parseable holdings table found for {source_url}")
        return max(candidates, key=len)

    def _parse_table(
        self,
        table,
        source_url: str,
        default_sector: Optional[str],
        etf_symbol: Optional[str],
    ) -> list[NormalizedHolding]:
        rows = table.find_all("tr")
        if len(rows) < 2:
            return []

        headers = self._extract_headers(rows[0])
        if not headers:
            return []
        mapping = self._map_headers(headers)
        if "ticker" not in mapping or "issuer_name" not in mapping:
            return []

        holdings: list[NormalizedHolding] = []
        for row in rows[1:]:
            values = self._extract_cells(row)
            if len(values) < len(headers):
                continue
            ticker = self._clean_text(values[mapping["ticker"]]).upper()
            issuer_name = self._clean_text(values[mapping["issuer_name"]])
            if not ticker or not issuer_name:
                continue
            sector = default_sector
            if "sector" in mapping:
                sector = self._clean_text(values[mapping["sector"]]) or default_sector
            cik = None
            if "cik" in mapping:
                cik = self._digits_only(values[mapping["cik"]]) or None
            weight = None
            if "weight" in mapping:
                weight = self._parse_weight(values[mapping["weight"]])

            holdings.append(
                NormalizedHolding(
                    ticker=ticker,
                    issuer_name=issuer_name,
                    sector=sector,
                    weight=weight,
                    cik=cik,
                    source_url=source_url,
                    etf_symbol=etf_symbol,
                )
            )
        return holdings

    def _extract_headers(self, row) -> list[str]:
        headers = [self._clean_text(cell.get_text(" ", strip=True)) for cell in row.find_all(["th", "td"])]
        return [header for header in headers if header]

    def _extract_cells(self, row) -> list[str]:
        return [cell.get_text(" ", strip=True) for cell in row.find_all(["td", "th"])]

    def _map_headers(self, headers: list[str]) -> dict[str, int]:
        mapped: dict[str, int] = {}
        normalized = [self._normalize_header(header) for header in headers]
        for canonical, aliases in self.COLUMN_ALIASES.items():
            for idx, header in enumerate(normalized):
                if header in aliases:
                    mapped[canonical] = idx
                    break
        return mapped

    def _normalize_header(self, value: str) -> str:
        return " ".join(value.strip().lower().replace("%", "").split())

    def _clean_text(self, value: str) -> str:
        return " ".join(value.strip().split())

    def _digits_only(self, value: str) -> str:
        return "".join(character for character in value if character.isdigit())

    def _parse_weight(self, value: str) -> Optional[float]:
        cleaned = value.replace("%", "").replace(",", "").strip()
        return coerce_float(cleaned)


class HoldingsExporter:
    def __init__(self, session: Optional[requests.Session] = None) -> None:
        self.session = session or requests.Session()
        self.parser = HoldingsHtmlParser()

    def fetch_html(self, url: str) -> str:
        response = self.session.get(url, headers=DEFAULT_HEADERS, timeout=30)
        response.raise_for_status()
        return response.text

    def fetch_bytes(self, url: str) -> bytes:
        response = self.session.get(url, headers=DEFAULT_HEADERS, timeout=30)
        response.raise_for_status()
        return response.content

    def export_ivv(
        self,
        output_path: str | Path,
        url: str,
        symbol: str = "IVV",
        product_id: str = "239726",
        fund_slug: str = "core-sp-500-etf",
    ) -> Path:
        download_url = ISHARES_DOWNLOAD_URL_TEMPLATE.format(
            product_id=product_id,
            fund_slug=fund_slug,
            symbol=symbol,
        )
        holdings: list[NormalizedHolding]
        try:
            csv_bytes = self.fetch_bytes(download_url)
            holdings = self._parse_ishares_csv(
                csv_bytes,
                source_url=download_url,
                fallback_source_url=url,
            )
        except Exception:
            html = self.fetch_html(url)
            holdings = self.parser.parse(html, source_url=url)
        return self._write(output_path, holdings)

    def export_spdr(self, output_path: str | Path, symbols: Optional[Iterable[str]] = None) -> Path:
        requested = list(symbols or SPDR_SECTOR_URLS.keys())
        holdings: list[NormalizedHolding] = []
        for symbol in requested:
            upper_symbol = symbol.upper()
            url = SPDR_SECTOR_URLS.get(upper_symbol)
            if not url:
                raise HoldingsExportError(f"unsupported SPDR sector ETF symbol: {symbol}")
            parsed = self._export_single_spdr_symbol(url=url, symbol=upper_symbol)
            holdings.extend(parsed)
        return self._write(output_path, holdings)

    def _export_single_spdr_symbol(self, url: str, symbol: str) -> list[NormalizedHolding]:
        for template in SPDR_HOLDINGS_URL_TEMPLATES:
            download_url = template.format(symbol=symbol.lower())
            try:
                workbook_bytes = self.fetch_bytes(download_url)
                parsed = self._parse_spdr_workbook(
                    workbook_bytes,
                    source_url=download_url,
                    symbol=symbol,
                )
                if parsed:
                    return parsed
            except Exception:
                continue

        html = self.fetch_html(url)
        return self.parser.parse(
            html,
            source_url=url,
            default_sector=symbol,
            etf_symbol=symbol,
        )

    def _parse_ishares_csv(
        self,
        csv_bytes: bytes,
        source_url: str,
        fallback_source_url: str,
    ) -> list[NormalizedHolding]:
        text = csv_bytes.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(StringIO(text)))
        header_row_index = None
        for index, row in enumerate(rows):
            normalized = [cell.strip().lower() for cell in row]
            if "ticker" in normalized and "name" in normalized and "sector" in normalized:
                header_row_index = index
                break
        if header_row_index is None:
            raise HoldingsExportError(f"iShares holdings CSV header not found in {source_url}")

        headers = [cell.strip() for cell in rows[header_row_index]]
        mapping = {header.lower(): idx for idx, header in enumerate(headers)}
        holdings: list[NormalizedHolding] = []
        for row in rows[header_row_index + 1 :]:
            if len(row) < len(headers):
                continue
            ticker = row[mapping["ticker"]].strip().upper()
            issuer_name = row[mapping["name"]].strip()
            if not ticker or not issuer_name:
                continue
            sector = row[mapping["sector"]].strip() if "sector" in mapping else None
            weight = None
            for candidate in ("weight (%)", "weight", "weight (%) "):
                if candidate in mapping:
                    weight = coerce_float(row[mapping[candidate]].replace("%", "").strip())
                    break
            cik = None
            if "cusip" in mapping:
                cik = None
            holdings.append(
                NormalizedHolding(
                    ticker=ticker,
                    issuer_name=issuer_name,
                    sector=sector,
                    weight=weight,
                    cik=cik,
                    source_url=fallback_source_url,
                )
            )
        if not holdings:
            raise HoldingsExportError(f"iShares holdings CSV contained no holdings rows from {source_url}")
        return holdings

    def _parse_spdr_workbook(
        self,
        workbook_bytes: bytes,
        source_url: str,
        symbol: str,
    ) -> list[NormalizedHolding]:
        workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header_row_index = None
        for index, row in enumerate(rows):
            cells = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            if "ticker" in cells and any(name in cells for name in ("company name", "name", "security name")):
                header_row_index = index
                break
        if header_row_index is None:
            raise HoldingsExportError(f"SPDR holdings workbook header not found in {source_url}")

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_row_index]]
        mapping = {header.lower(): idx for idx, header in enumerate(headers)}
        name_key = next(
            (candidate for candidate in ("company name", "name", "security name") if candidate in mapping),
            None,
        )
        if name_key is None:
            raise HoldingsExportError(f"SPDR holdings workbook name column not found in {source_url}")

        holdings: list[NormalizedHolding] = []
        for row in rows[header_row_index + 1 :]:
            if not row or len(row) < len(headers):
                continue
            ticker = str(row[mapping["ticker"]] or "").strip().upper()
            issuer_name = str(row[mapping[name_key]] or "").strip()
            if not ticker or not issuer_name:
                continue
            weight = None
            for candidate in ("weight", "index weight", "fund weight", "% weight"):
                if candidate in mapping:
                    raw = row[mapping[candidate]]
                    if isinstance(raw, str):
                        weight = coerce_float(raw.replace("%", "").strip())
                    else:
                        weight = coerce_float(raw)
                    break
            sector = None
            for candidate in ("sector", "gics sector"):
                if candidate in mapping:
                    sector_value = row[mapping[candidate]]
                    sector = str(sector_value).strip() if sector_value is not None else None
                    break
            holdings.append(
                NormalizedHolding(
                    ticker=ticker,
                    issuer_name=issuer_name,
                    sector=sector or symbol,
                    weight=weight,
                    source_url=source_url,
                    etf_symbol=symbol,
                )
            )
        if not holdings:
            raise HoldingsExportError(f"SPDR holdings workbook contained no holdings rows from {source_url}")
        return holdings

    def _write(self, output_path: str | Path, holdings: list[NormalizedHolding]) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        rows = [holding.as_record() for holding in holdings]
        path.write_text(json.dumps(rows, indent=2, ensure_ascii=True), encoding="utf-8")
        return path
